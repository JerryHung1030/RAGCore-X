import openpyxl
import json

def remove_empty_evidence_predictions(output_str):
    """
    將輸入的 JSON 字串中, 'predictions' 裡 evidence 為空陣列的物件移除後回傳新的 JSON 字串。
    """
    if not output_str:
        return output_str
    
    # 嘗試解析成 dict
    try:
        data = json.loads(output_str)
    except json.JSONDecodeError:
        # 若非合法 JSON, 原樣回傳
        print("JSON 解析失敗, 跳過不處理。")
        return output_str
    
    # 若有 predictions 欄位，且為 list：
    if 'predictions' in data and isinstance(data['predictions'], list):
        filtered = []
        for pred in data['predictions']:
            # 只保留 evidence 不為空的物件
            if pred.get('evidence'):
                filtered.append(pred)
        data['predictions'] = filtered
    
    # 轉回 JSON 字串 (為了方便閱讀，indent=4；也可不縮排)
    return json.dumps(data, ensure_ascii=False, indent=4)

def main():
    excel_file = "fraud_evaluation_data_format.xlsx"
    sheet_name = "reverse_4omini_k15_low"
    
    # 載入工作簿 (不使用 data_only=True，才能原生讀寫)
    wb = openpyxl.load_workbook(excel_file)
    
    # 取得指定工作表
    if sheet_name not in wb.sheetnames:
        print(f"工作表 {sheet_name} 不存在，請確認名稱。")
        return
    ws = wb[sheet_name]
    
    # 逐列讀取與修改
    # 假設第 1 列為標題，從第 2 列開始處理至資料結束
    start_row = 2
    max_row = ws.max_row
    
    input_col = 1   # A 欄
    output_col = 2  # B 欄
    
    for row_idx in range(start_row, max_row + 1):
        # 讀取 input 欄資料 (若需要的話)
        input_cell = ws.cell(row=row_idx, column=input_col)
        input_value = input_cell.value
        
        # 讀取 output 欄資料
        output_cell = ws.cell(row=row_idx, column=output_col)
        output_value = output_cell.value
        
        if output_value:
            # 對 output_value 做移除 empty evidence 物件的動作
            cleaned_json = remove_empty_evidence_predictions(output_value)
            
            # 將結果寫回到原本同一個儲存格
            output_cell.value = cleaned_json
        else:
            # 可能是空或 None，就略過或自行決定處理邏輯
            pass
    
    # 將修改後的工作表直接儲存 (保留合併儲存格、格式等)
    wb.save("fraud_evaluation_data_format.xlsx")
    print("處理完成，已保留原先的合併儲存格格式。")

if __name__ == "__main__":
    main()
